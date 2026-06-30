from openpyxl import Workbook
from openpyxl.styles import Font


def generate_report(
    file_path,
    username
):
    wb = Workbook()
    ws = wb.active

    ws.title = "SentinelX AI Report"

    ws["A1"] = "SentinelX AI Report"
    ws["A1"].font = Font(
        bold=True,
        size=14
    )

    ws["A3"] = "User Name"
    ws["B3"] = username

    ws["A4"] = "Report Type"
    ws["B4"] = "SentinelX AI Security Report"

    ws["A5"] = "Status"
    ws["B5"] = "Generated Successfully"

    wb.save(file_path)